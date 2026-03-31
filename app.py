"""
PE Value Creation Platform

Upload → Ingest → Auto-Suggest Assumptions → Tweak → Project → Analyze → Excel
Run with: streamlit run app.py
"""

import io
import json
import os
import tempfile
from datetime import date

import pandas as pd
import streamlit as st

from core.ingest import ingest_file
from core.result import NormalizedResult
from analysis.engine import run_analysis
from analysis.types import Favorability, Severity
from analysis.excel_export import export_to_excel
from modeling.types import (
    AssumptionSet, RevenueAssumptions, CostAssumptions, CostLineAssumption,
    WorkingCapitalAssumptions, CapExAssumptions, DebtAssumptions,
    TaxAssumptions, ExitAssumptions, Initiative,
)
from modeling.auto_suggest import suggest_assumptions
from modeling.engine import run_model


st.set_page_config(page_title="PE Value Creation Platform", page_icon="📊", layout="wide")

st.markdown("""
<style>
    .flag-critical { border-left: 4px solid #ef4444; background: #fef2f2; padding: 8px 12px; border-radius: 6px; margin: 4px 0; }
    .flag-warning { border-left: 4px solid #f59e0b; background: #fffbeb; padding: 8px 12px; border-radius: 6px; margin: 4px 0; }
    .flag-info { border-left: 4px solid #3b82f6; background: #eff6ff; padding: 8px 12px; border-radius: 6px; margin: 4px 0; }
    .bridge-positive { color: #10b981; font-weight: 600; }
    .bridge-negative { color: #ef4444; font-weight: 600; }
</style>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("## 📊 PE Value Creation")
    st.divider()

    mode = st.radio("Mode", ["Demo Data (all types)", "Upload my own file"])

    DEMO_FILES = {
        "P&L — Clean CSV": ("data/sample_pl.csv", "Meridian Software", "B2B SaaS"),
        "P&L — QuickBooks Export": ("data/test/quickbooks_export.csv", "Acme Corp", "B2B SaaS"),
        "P&L — Manufacturing Excel": ("data/test/manufacturing_pl.xlsx", "Atlas Manufacturing", "Manufacturing"),
        "P&L — Messy $000s Workbook": ("data/test/messy_workbook.xlsx", "Meridian Software", "B2B SaaS"),
        "Balance Sheet": ("data/test/balance_sheet_clean.csv", "Meridian Software", "B2B SaaS"),
        "Cash Flow Statement": ("data/test/cash_flow_clean.csv", "Meridian Software", "B2B SaaS"),
        "Working Capital / AR Aging": ("data/test/working_capital_clean.csv", "Meridian Software", "B2B SaaS"),
        "Revenue Detail (by product)": ("data/test/revenue_detail_clean.csv", "Meridian Software", "B2B SaaS"),
        "KPI / Operational Metrics": ("data/test/kpi_operational_clean.csv", "Meridian Software", "B2B SaaS"),
    }

    uploaded_file = None
    selected_files = {}

    if mode == "Upload my own file":
        uploaded_file = st.file_uploader("Upload file", type=["csv", "xlsx", "xls", "pdf"])
        company_name = st.text_input("Company Name", "")
        business_type = st.selectbox("Business Type", ["B2B SaaS", "Services", "Manufacturing", "Distribution", "Other"])
    else:
        st.caption("Select files:")
        for label, (path, company, btype) in DEMO_FILES.items():
            if st.checkbox(label, value=(label in [
                "P&L — Clean CSV", "Balance Sheet", "Cash Flow Statement",
                "Working Capital / AR Aging", "Revenue Detail (by product)",
                "KPI / Operational Metrics",
            ])):
                selected_files[label] = (path, company, btype)

    st.divider()
    ingest_btn = st.button("📂 Load Data", type="primary", use_container_width=True)


# ═══════════════════════════════════════════════════════════════════════════
# INGEST (only when Load Data clicked)
# ═══════════════════════════════════════════════════════════════════════════

st.markdown("# PE Value Creation Platform")

# Store ingested data in session state so it persists
if ingest_btn:
    ingested = {}
    if mode == "Upload my own file":
        if uploaded_file:
            suffix = os.path.splitext(uploaded_file.name)[1]
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                tmp.write(uploaded_file.read())
            try:
                result = ingest_file(tmp.name, company_name=company_name, business_type=business_type)
                ingested[result.doc_type] = result
            except Exception as e:
                st.error(f"Failed: {e}")
    else:
        for label, (path, company, btype) in selected_files.items():
            try:
                result = ingest_file(path, company_name=company, business_type=btype)
                ingested[result.doc_type] = result
            except Exception as e:
                st.warning(f"Failed: {label}: {e}")

    if ingested:
        st.session_state["ingested"] = ingested
        st.session_state["suggested"] = suggest_assumptions(ingested)

# Check if we have data
if "ingested" not in st.session_state:
    st.info("Select data in the sidebar and click **Load Data**.")
    st.stop()

ingested = st.session_state["ingested"]
suggested = st.session_state["suggested"]

# Show ingestion summary
st.header("① Data Loaded")
cols = st.columns(min(len(ingested), 6))
for i, (doc_type, result) in enumerate(ingested.items()):
    with cols[i % len(cols)]:
        method = "⚡" if not result.used_ai else "🤖"
        st.metric(result.doc_type_name.split("/")[0].strip()[:18], f"{len(result.df)} rows", method)

st.divider()

# ═══════════════════════════════════════════════════════════════════════════
# ASSUMPTIONS (editable, auto-updates analysis)
# ═══════════════════════════════════════════════════════════════════════════

st.header("② Assumptions")
st.caption("Auto-generated from your historical data. Edit any value to update projections.")

ac1, ac2, ac3 = st.columns(3)

with ac1:
    st.subheader("Revenue")
    rev_growth = st.number_input(
        "Monthly Growth Rate %",
        value=suggested.revenue.growth_rate_pct or 0.0,
        step=0.5, format="%.1f",
        help="Median MoM growth from last 6 months",
    )
    proj_months = st.number_input("Projection Months", value=12, min_value=1, max_value=60, step=6)

with ac2:
    st.subheader("Costs (% of Revenue)")
    # Find suggested values
    def _get_suggested_pct(line_item):
        for cl in suggested.costs.lines:
            if cl.line_item == line_item:
                return cl.pct_of_revenue or 0.0
        return 0.0

    cogs_pct = st.number_input("COGS %", value=_get_suggested_pct("cogs"), step=0.5, format="%.1f")
    sm_pct = st.number_input("S&M %", value=_get_suggested_pct("sales_marketing"), step=0.5, format="%.1f")
    rd_pct = st.number_input("R&D %", value=_get_suggested_pct("rd"), step=0.5, format="%.1f")
    ga_pct = st.number_input("G&A %", value=_get_suggested_pct("ga"), step=0.5, format="%.1f")

with ac3:
    st.subheader("Working Capital")
    dso = st.number_input("Target DSO (days)", value=suggested.working_capital.target_dso or 40.0, step=1.0, format="%.0f")
    dpo = st.number_input("Target DPO (days)", value=suggested.working_capital.target_dpo or 30.0, step=1.0, format="%.0f")

    st.subheader("Exit")
    exit_multiple = st.number_input("Exit Multiple (x)", value=10.0, step=0.5, format="%.1f")
    entry_equity = st.number_input("Entry Equity ($M)", value=10.0, step=1.0, format="%.1f")
    exit_year = st.number_input("Exit Year", value=5, min_value=1, max_value=10, step=1)

# Show implied EBITDA margin
implied_margin = 100 - cogs_pct - sm_pct - rd_pct - ga_pct
margin_color = "green" if implied_margin > 10 else ("orange" if implied_margin > 0 else "red")
st.markdown(f"**Implied EBITDA Margin: :{margin_color}[{implied_margin:.1f}%]**")

st.divider()

# ═══════════════════════════════════════════════════════════════════════════
# BUILD ASSUMPTIONS + RUN MODEL
# ═══════════════════════════════════════════════════════════════════════════

# Build AssumptionSet from UI inputs
assumptions = AssumptionSet(
    name="Base",
    projection_months=proj_months,
    revenue=RevenueAssumptions(method="growth_rate", growth_rate_pct=rev_growth),
    costs=CostAssumptions(lines=[
        CostLineAssumption("cogs", method="pct_of_revenue", pct_of_revenue=cogs_pct),
        CostLineAssumption("sales_marketing", method="pct_of_revenue", pct_of_revenue=sm_pct),
        CostLineAssumption("rd", method="pct_of_revenue", pct_of_revenue=rd_pct),
        CostLineAssumption("ga", method="pct_of_revenue", pct_of_revenue=ga_pct),
    ]),
    working_capital=WorkingCapitalAssumptions(target_dso=dso, target_dpo=dpo),
    capex=suggested.capex,
    debt=suggested.debt,
    tax=suggested.tax,
    exit_=ExitAssumptions(
        exit_year=exit_year,
        exit_multiple=exit_multiple,
        entry_equity=entry_equity * 1_000_000,
    ),
)

# Run model (projections + analysis + returns)
model = run_model(ingested, assumptions)
analysis = model.analysis

# ═══════════════════════════════════════════════════════════════════════════
# RESULTS
# ═══════════════════════════════════════════════════════════════════════════

st.header("③ Analysis + Projections")

# Key metrics row
if model.returns and model.returns.exit_equity:
    r = model.returns
    mc1, mc2, mc3, mc4, mc5 = st.columns(5)
    with mc1:
        st.metric("MOIC", f"{r.moic:.2f}x" if r.moic else "N/A")
    with mc2:
        st.metric("IRR", f"{r.irr:.1%}" if r.irr else "N/A")
    with mc3:
        st.metric("Exit EV", f"${r.exit_ev/1e6:.1f}M" if r.exit_ev else "N/A")
    with mc4:
        st.metric("Exit Equity", f"${r.exit_equity/1e6:.1f}M" if r.exit_equity else "N/A")
    with mc5:
        if analysis.ltm:
            st.metric("Rule of 40", f"{analysis.ltm.rule_of_40:.0f}" if analysis.ltm.rule_of_40 else "N/A")

if analysis.ltm:
    lc1, lc2, lc3, lc4 = st.columns(4)
    with lc1:
        st.metric("LTM Revenue", f"${analysis.ltm.ltm_revenue/1e6:.1f}M" if analysis.ltm.ltm_revenue else "N/A")
    with lc2:
        st.metric("LTM EBITDA", f"${analysis.ltm.ltm_ebitda/1e6:.1f}M" if analysis.ltm.ltm_ebitda else "N/A")
    with lc3:
        st.metric("LTM EBITDA Margin", f"{analysis.ltm.ltm_ebitda_margin_pct:.1f}%" if analysis.ltm.ltm_ebitda_margin_pct else "N/A")
    with lc4:
        st.metric("Modules Run", len(analysis.modules_run))

st.divider()

# ── Tabs ──
tab_names = []
if analysis.ebitda_bridges: tab_names.append("EBITDA Bridge")
if analysis.margins: tab_names.append("Margins & Growth")
if analysis.variance: tab_names.append("Variance Analysis")
if analysis.working_capital: tab_names.append("Working Capital")
if analysis.fcf: tab_names.append("FCF & Leverage")
if analysis.ltm: tab_names.append("LTM & Rule of 40")
if analysis.revenue_analytics: tab_names.append("Revenue Analytics")
if analysis.trends: tab_names.append("Trend Flags")
tab_names.append("Projected P&L")

if not tab_names:
    st.info("No analysis results.")
    st.stop()

tabs = st.tabs(tab_names)
tab_idx = 0

# ── EBITDA Bridge ──
if analysis.ebitda_bridges:
    with tabs[tab_idx]:
        eb = analysis.ebitda_bridges
        bridge_cols = st.columns(3 if eb.vs_prior_year else (2 if eb.vs_budget else 1))

        def render_bridge(bridge, col):
            if bridge is None: return
            with col:
                st.subheader(bridge.label)
                st.caption(f"{bridge.base_period} → {bridge.current_period}")
                st.markdown(f"**Starting EBITDA:** ${bridge.base_ebitda:,.0f}")
                for comp in bridge.components:
                    color = "bridge-positive" if comp.value >= 0 else "bridge-negative"
                    st.markdown(f'<div style="display:flex;justify-content:space-between;padding:4px 0;"><span>{comp.name}</span><span class="{color}">${comp.value:+,.0f}</span></div>', unsafe_allow_html=True)
                st.markdown("---")
                st.markdown(f"**Ending EBITDA:** ${bridge.current_ebitda:,.0f}")
                st.markdown(f"**Total Change:** ${bridge.total_change:+,.0f}")
                st.caption("✅ Verified" if bridge.is_verified else f"⚠️ Delta: ${bridge.verification_delta:.2f}")

        render_bridge(eb.mom, bridge_cols[0])
        if eb.vs_budget and len(bridge_cols) > 1: render_bridge(eb.vs_budget, bridge_cols[1])
        if eb.vs_prior_year and len(bridge_cols) > 2: render_bridge(eb.vs_prior_year, bridge_cols[2])
    tab_idx += 1

# ── Margins ──
if analysis.margins:
    with tabs[tab_idx]:
        m = analysis.margins
        latest = m.periods[-1]
        mc1, mc2, mc3, mc4, mc5 = st.columns(5)
        with mc1: st.metric("Gross Margin", f"{latest.gross_margin_pct:.1f}%" if latest.gross_margin_pct else "N/A")
        with mc2: st.metric("EBITDA Margin", f"{latest.ebitda_margin_pct:.1f}%" if latest.ebitda_margin_pct else "N/A")
        with mc3: st.metric("Rev Growth MoM", f"{latest.revenue_growth_mom:+.1f}%" if latest.revenue_growth_mom else "N/A")
        with mc4: st.metric("Rev Growth YoY", f"{latest.revenue_growth_yoy:+.1f}%" if latest.revenue_growth_yoy else "N/A")
        with mc5: st.metric("OpEx % Rev", f"{latest.opex_pct_revenue:.1f}%" if latest.opex_pct_revenue else "N/A")
        if not m.as_dataframe.empty:
            display = m.as_dataframe.copy()
            for col in display.columns:
                if col != "period" and display[col].dtype in ['float64', 'int64']:
                    display[col] = display[col].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "")
            st.dataframe(display, use_container_width=True, hide_index=True)
    tab_idx += 1

# ── Variance ──
if analysis.variance:
    with tabs[tab_idx]:
        latest_var = analysis.variance.periods[-1]
        st.subheader(f"Variance — {latest_var.period}")
        var_data = latest_var.vs_prior_month
        if var_data:
            rows = []
            for v in var_data:
                fav = "🟢" if v.favorable == Favorability.FAVORABLE else ("🔴" if v.favorable == Favorability.UNFAVORABLE else "⚪")
                rows.append({"": fav, "Line": v.line_item.replace("_"," ").title(), "Actual": f"${v.actual:,.0f}", "Prior": f"${v.comparator:,.0f}", "$ Chg": f"${v.dollar_change:+,.0f}", "% Chg": f"{v.pct_change:+.1f}%" if v.pct_change else "N/A"})
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
    tab_idx += 1

# ── Working Capital ──
if analysis.working_capital:
    with tabs[tab_idx]:
        wc = analysis.working_capital
        latest_wc = wc.periods[-1]
        wc1, wc2, wc3, wc4 = st.columns(4)
        with wc1: st.metric("DSO", f"{latest_wc.dso:.0f} days" if latest_wc.dso else "N/A")
        with wc2: st.metric("DPO", f"{latest_wc.dpo:.0f} days" if latest_wc.dpo else "N/A")
        with wc3: st.metric("CCC", f"{latest_wc.ccc:.0f} days" if latest_wc.ccc else "N/A")
        with wc4:
            if latest_wc.wc_change is not None:
                st.metric("WC Change", f"${latest_wc.wc_change:+,.0f}")
        wc_rows = [{"Period": p.period, "DSO": f"{p.dso:.0f}" if p.dso else "", "DPO": f"{p.dpo:.0f}" if p.dpo else "", "CCC": f"{p.ccc:.0f}" if p.ccc else "", "WC Chg": f"${p.wc_change:+,.0f}" if p.wc_change is not None else ""} for p in wc.periods]
        st.dataframe(pd.DataFrame(wc_rows), use_container_width=True, hide_index=True)
    tab_idx += 1

# ── FCF ──
if analysis.fcf:
    with tabs[tab_idx]:
        latest_fcf = analysis.fcf.periods[-1]
        fc1, fc2, fc3 = st.columns(3)
        with fc1: st.metric("FCF", f"${latest_fcf.free_cash_flow:,.0f}" if latest_fcf.free_cash_flow else "N/A")
        with fc2: st.metric("Cash Conv.", f"{latest_fcf.cash_conversion_ratio:.1%}" if latest_fcf.cash_conversion_ratio else "N/A")
        with fc3: st.metric("ND/EBITDA", f"{latest_fcf.net_debt_to_ltm_ebitda:.1f}x" if latest_fcf.net_debt_to_ltm_ebitda else "N/A")
        fcf_rows = [{"Period": p.period, "FCF": f"${p.free_cash_flow:,.0f}" if p.free_cash_flow else "", "Cash Conv": f"{p.cash_conversion_ratio:.1%}" if p.cash_conversion_ratio else "", "Net Debt": f"${p.net_debt:,.0f}" if p.net_debt is not None else "", "ND/EBITDA": f"{p.net_debt_to_ltm_ebitda:.1f}x" if p.net_debt_to_ltm_ebitda else ""} for p in analysis.fcf.periods]
        st.dataframe(pd.DataFrame(fcf_rows), use_container_width=True, hide_index=True)
    tab_idx += 1

# ── LTM ──
if analysis.ltm:
    with tabs[tab_idx]:
        ltm = analysis.ltm
        st.subheader(f"Last Twelve Months — as of {ltm.as_of_period}")
        lc1, lc2, lc3, lc4 = st.columns(4)
        with lc1: st.metric("LTM Revenue", f"${ltm.ltm_revenue:,.0f}" if ltm.ltm_revenue else "N/A")
        with lc2: st.metric("LTM EBITDA", f"${ltm.ltm_ebitda:,.0f}" if ltm.ltm_ebitda else "N/A")
        with lc3: st.metric("LTM Margin", f"{ltm.ltm_ebitda_margin_pct:.1f}%" if ltm.ltm_ebitda_margin_pct else "N/A")
        with lc4: st.metric("Rev Growth YoY", f"{ltm.ltm_revenue_growth_yoy:+.1f}%" if ltm.ltm_revenue_growth_yoy else "N/A")
        if ltm.rule_of_40 is not None:
            r40_color = "green" if ltm.rule_of_40 >= 40 else "red"
            st.markdown(f"**Rule of 40: :{r40_color}[{ltm.rule_of_40:.1f}]**")
    tab_idx += 1

# ── Revenue Analytics ──
if analysis.revenue_analytics:
    with tabs[tab_idx]:
        ra = analysis.revenue_analytics
        if ra.concentration:
            c = ra.concentration[-1]
            rc1, rc2, rc3 = st.columns(3)
            with rc1: st.metric("Top 1", f"{c.top1_pct:.1f}%" if c.top1_pct else "N/A")
            with rc2: st.metric(f"{c.dimension.title()}s", c.count)
            with rc3: st.metric("HHI", f"{c.herfindahl:.3f}")
        if ra.price_volume:
            st.subheader("Price / Volume / Mix")
            pv_rows = [{"Period": pv.period, "Price": f"${pv.price_effect:+,.0f}", "Volume": f"${pv.volume_effect:+,.0f}", "Mix": f"${pv.mix_effect:+,.0f}", "Total": f"${pv.total_change:+,.0f}", "✓": "✅" if pv.is_verified else "❌"} for pv in ra.price_volume]
            st.dataframe(pd.DataFrame(pv_rows), use_container_width=True, hide_index=True)
        if ra.kpi_trends:
            st.subheader("KPI Trends")
            for name, series in ra.kpi_trends.items():
                kdf = pd.DataFrame(series, columns=["Period", name.replace("_"," ").title()])
                st.line_chart(kdf.set_index("Period"))
    tab_idx += 1

# ── Trends ──
if analysis.trends:
    with tabs[tab_idx]:
        flags = analysis.trends.flags
        if not flags:
            st.success("No trend flags — all metrics normal.")
        else:
            crit = sum(1 for f in flags if f.severity == Severity.CRITICAL)
            warn = sum(1 for f in flags if f.severity == Severity.WARNING)
            sc1, sc2, sc3 = st.columns(3)
            with sc1: st.metric("🔴 Critical", crit)
            with sc2: st.metric("🟡 Warning", warn)
            with sc3: st.metric("🔵 Info", len(flags) - crit - warn)
            for sev, css in [(Severity.CRITICAL, "flag-critical"), (Severity.WARNING, "flag-warning"), (Severity.INFO, "flag-info")]:
                for f in [f for f in flags if f.severity == sev]:
                    st.markdown(f'<div class="{css}"><strong>{f.metric.replace("_"," ").title()}</strong> — {f.flag_type.value.replace("_"," ").title()}<br/><span style="color:#666">{f.detail}</span></div>', unsafe_allow_html=True)
    tab_idx += 1

# ── Projected P&L ──
with tabs[tab_idx]:
    st.subheader("Projected P&L (Historical + Forecast)")
    if "income_statement" in model.combined_data:
        is_df = model.combined_data["income_statement"].df
        display = is_df.copy()
        pcol = "period" if "period" in display.columns else "month"
        for col in display.columns:
            if col not in [pcol, "_is_projected"] and pd.api.types.is_numeric_dtype(display[col]):
                display[col] = display[col].apply(lambda x: f"${x:,.0f}" if pd.notna(x) and abs(x) >= 100 else (str(x) if pd.notna(x) else ""))
        display["Type"] = display["_is_projected"].apply(lambda x: "📈 Projected" if x else "📊 Actual")
        display = display.drop(columns=["_is_projected"])
        cols_order = [pcol, "Type"] + [c for c in display.columns if c not in [pcol, "Type"]]
        st.dataframe(display[cols_order], use_container_width=True, hide_index=True, height=500)

st.divider()

# ═══════════════════════════════════════════════════════════════════════════
# EXPORT
# ═══════════════════════════════════════════════════════════════════════════

st.header("④ Export")

# Excel preview + download
from openpyxl import load_workbook

excel_buffer = io.BytesIO()
company = next(iter(ingested.values())).company_name if ingested else ""
export_to_excel(analysis, excel_buffer, ingested=model.combined_data, company_name=company)
excel_bytes = excel_buffer.getvalue()

preview_wb = load_workbook(io.BytesIO(excel_bytes))
preview_tabs = st.tabs(preview_wb.sheetnames)
for i, sheet_name in enumerate(preview_wb.sheetnames):
    with preview_tabs[i]:
        ws = preview_wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(v) if v is not None else "" for v in row])
        if rows:
            num_cols = max(len(r) for r in rows)
            col_names = [f"Col {j+1}" for j in range(num_cols)]
            padded = [r + [""] * (num_cols - len(r)) for r in rows]
            st.dataframe(pd.DataFrame(padded, columns=col_names), use_container_width=True, hide_index=True, height=300)

st.download_button(
    "📥 Download Excel Workbook",
    excel_bytes,
    file_name=f"analysis_{date.today().isoformat()}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)

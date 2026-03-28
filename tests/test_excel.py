"""Comprehensive Excel export + math verification tests."""

import io
import sys
import pandas as pd
from openpyxl import load_workbook

from core.ingest import ingest_file
from analysis.engine import run_analysis
from analysis.excel_export import export_to_excel

PASS = 0
FAIL = 0

def check(name, condition, detail=""):
    global PASS, FAIL
    if condition:
        PASS += 1
    else:
        FAIL += 1
        print(f"  FAIL: {name} — {detail}")

def approx(a, b, tol=0.5):
    if a is None or b is None:
        return False
    return abs(a - b) < tol


print("=" * 70)
print("  COMPREHENSIVE EXCEL + ANALYSIS VERIFICATION")
print("=" * 70)

# ── 1. Full pipeline ──
print("\n── Test 1: Full pipeline (6 data types) ──")
data = {}
for doc_type, path in [
    ("income_statement", "data/sample_pl.csv"),
    ("balance_sheet", "data/test/balance_sheet_clean.csv"),
    ("cash_flow", "data/test/cash_flow_clean.csv"),
    ("working_capital", "data/test/working_capital_clean.csv"),
    ("revenue_detail", "data/test/revenue_detail_clean.csv"),
    ("kpi_operational", "data/test/kpi_operational_clean.csv"),
]:
    data[doc_type] = ingest_file(path)

analysis = run_analysis(data)
check("7 modules", len(analysis.modules_run) == 7, f"got {len(analysis.modules_run)}")
check("0 warnings", len(analysis.warnings) == 0, str(analysis.warnings))

# ── 2. Generate Excel ──
print("\n── Test 2: Generate Excel ──")
buf = io.BytesIO()
export_to_excel(analysis, buf, ingested=data, company_name="Meridian Software")
excel_bytes = buf.getvalue()
check("Excel not empty", len(excel_bytes) > 1000, f"{len(excel_bytes)} bytes")

# ── 3. Sheet structure ──
print("\n── Test 3: Sheet structure ──")
wb = load_workbook(io.BytesIO(excel_bytes))
expected = ["Summary", "EBITDA Bridges", "Variance Analysis", "Margins & Growth",
            "Working Capital", "FCF & Leverage", "Revenue Analytics", "Trend Flags", "Raw Data"]
check("9 sheets", len(wb.sheetnames) == 9, str(wb.sheetnames))
for s in expected:
    check(f"Sheet '{s}' exists", s in wb.sheetnames)

# ── 4. EBITDA Bridge math ──
print("\n── Test 4: EBITDA Bridge cell values ──")
ws = wb["EBITDA Bridges"]
base = ws.cell(3, 2).value
rev = ws.cell(4, 2).value
cogs = ws.cell(5, 2).value
sm = ws.cell(6, 2).value
rd = ws.cell(7, 2).value
ga = ws.cell(8, 2).value
end = ws.cell(9, 2).value
total = ws.cell(10, 2).value

check("Base EBITDA = 207000", base == 207000, str(base))
check("Revenue impact = 300000", rev == 300000, str(rev))
check("COGS impact = -64000", cogs == -64000, str(cogs))
check("S&M impact = -74000", sm == -74000, str(sm))
check("R&D impact = -56200", rd == -56200, str(rd))
check("G&A impact = -23800", ga == -23800, str(ga))
check("Ending EBITDA = 289000", end == 289000, str(end))
check("Total change = 82000", total == 82000, str(total))

# Components sum = total
comp_sum = rev + cogs + sm + rd + ga
check("Components sum to total", abs(comp_sum - total) < 0.01, f"{comp_sum} vs {total}")
# Total = end - base
check("Total = end - base", abs((end - base) - total) < 0.01)
# Verified flag
verified = ws.cell(11, 2).value
check("Bridge verified=Yes", verified == "Yes", str(verified))

# ── 5. Budget bridge ──
print("\n── Test 5: Budget bridge ──")
# Find budget bridge section (after MoM bridge ends around row 12)
budget_base = None
for r in range(12, 30):
    val = ws.cell(r, 1).value
    if val and "Budget" in str(val) and "Starting" in str(val):
        budget_base = ws.cell(r, 2).value
        break
check("Budget bridge exists", budget_base is not None, "Not found")
if budget_base:
    check("Budget base = 340000", budget_base == 340000, str(budget_base))

# ── 6. Variance analysis ──
print("\n── Test 6: Variance analysis ──")
ws_v = wb["Variance Analysis"]
check("Rev actual = 3400000", ws_v.cell(3, 2).value == 3400000)
check("Rev prior = 3100000", ws_v.cell(3, 3).value == 3100000)
check("Rev change = 300000", ws_v.cell(3, 4).value == 300000)
# COGS: actual=901000, prior=837000, change=64000
check("COGS actual = 901000", ws_v.cell(4, 2).value == 901000)
check("COGS change = 64000", ws_v.cell(4, 4).value == 64000)
# Revenue variance math: actual - prior = change
rev_a = ws_v.cell(3, 2).value
rev_p = ws_v.cell(3, 3).value
rev_c = ws_v.cell(3, 4).value
check("Variance math: actual-prior=change", rev_a - rev_p == rev_c)

# Budget section
# Find budget rev row
for r in range(11, 20):
    if ws_v.cell(r, 1).value == "Revenue":
        bud_actual = ws_v.cell(r, 2).value
        bud_comp = ws_v.cell(r, 3).value
        bud_change = ws_v.cell(r, 4).value
        check("Budget rev actual = 3400000", bud_actual == 3400000)
        check("Budget rev comp = 3600000", bud_comp == 3600000)
        check("Budget rev change = -200000", bud_change == -200000)
        break

# ── 7. Margins ──
print("\n── Test 7: Margins ──")
ws_m = wb["Margins & Growth"]
check("Row 2 = Gross Margin %", ws_m.cell(2, 1).value == "Gross Margin %")
check("Row 3 = EBITDA Margin %", ws_m.cell(3, 1).value == "EBITDA Margin %")
# Last column (14 = header + 13 periods) should be ~73.5 gross margin
gm_last = ws_m.cell(2, 14).value
check("GM latest ~73.5", approx(gm_last, 73.5), str(gm_last))
em_last = ws_m.cell(3, 14).value
check("EBITDA margin latest ~8.5", approx(em_last, 8.5), str(em_last))

# ── 8. Working Capital ──
print("\n── Test 8: Working Capital ──")
ws_wc = wb["Working Capital"]
check("Row 2 = DSO", ws_wc.cell(2, 1).value == "DSO (days)")
check("Row 5 = Cash Conversion Cycle", ws_wc.cell(5, 1).value == "Cash Conversion Cycle")
# Last period DSO should be 38
dso = ws_wc.cell(2, 7).value  # col 7 = 6th period
check("DSO latest = 38", dso == 38, str(dso))

# ── 9. FCF ──
print("\n── Test 9: FCF ──")
ws_f = wb["FCF & Leverage"]
check("Row 2 = FCF", "Free Cash Flow" in str(ws_f.cell(2, 1).value))
fcf = ws_f.cell(2, 7).value
check("FCF is positive", fcf is not None and fcf > 0, str(fcf))

# ── 10. Preview rendering (no duplicate column crash) ──
print("\n── Test 10: Preview rendering ──")
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(v) if v is not None else "" for v in row])
    if rows:
        num_cols = max(len(r) for r in rows)
        col_names = [f"Col {j+1}" for j in range(num_cols)]
        padded = [r + [""] * (num_cols - len(r)) for r in rows]
        preview_df = pd.DataFrame(padded, columns=col_names)
        check(f"Preview {sheet_name}", len(preview_df) > 0)

# ── 11. Minimal data (P&L only) ──
print("\n── Test 11: P&L only export ──")
data_min = {"income_statement": ingest_file("data/sample_pl.csv")}
analysis_min = run_analysis(data_min)
buf_min = io.BytesIO()
export_to_excel(analysis_min, buf_min, ingested=data_min)
wb_min = load_workbook(io.BytesIO(buf_min.getvalue()))
check("Min has Summary", "Summary" in wb_min.sheetnames)
check("Min has EBITDA", "EBITDA Bridges" in wb_min.sheetnames)
check("Min has Margins", "Margins & Growth" in wb_min.sheetnames)
# Should NOT have WC or FCF tabs (no data)
check("Min no WC tab", "Working Capital" not in wb_min.sheetnames)
check("Min no FCF tab", "FCF & Leverage" not in wb_min.sheetnames)

# ── 12. Empty export ──
print("\n── Test 12: Empty export ──")
analysis_empty = run_analysis({})
buf_empty = io.BytesIO()
export_to_excel(analysis_empty, buf_empty)
wb_empty = load_workbook(io.BytesIO(buf_empty.getvalue()))
check("Empty has Summary", "Summary" in wb_empty.sheetnames)

# ── Results ──
print("\n" + "=" * 70)
total = PASS + FAIL
print(f"  RESULTS: {PASS} passed, {FAIL} failed out of {total} tests")
print("=" * 70)

if FAIL > 0:
    sys.exit(1)

"""
Excel export — renders pre-computed AnalysisResult into a formatted workbook.

All math is already done. This is purely a rendering step.
No Excel formulas. Every cell is a pre-computed value.

Style: Clean, simple, professional.
- White background, black text
- Dark blue headers with white text
- Thin borders on data cells
- Consistent across every tab
"""

from datetime import date
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter

from analysis.types import (
    AnalysisResult, EBITDABridge, Favorability, Severity,
)
from core.result import NormalizedResult


# ── Styles (consistent everywhere) ──

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

SECTION_FONT = Font(bold=True, size=11, color="1F4E79")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

LABEL_FONT = Font(bold=True, size=10)
DATA_FONT = Font(size=10)
BOLD_FONT = Font(bold=True, size=10)

THIN_BORDER = Border(
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
)

FMT_CURRENCY = '$#,##0'
FMT_CURRENCY_SIGN = '$#,##0;($#,##0)'
FMT_PCT = '0.0%'
FMT_PCT_RAW = '0.0'  # for values already in percentage form (e.g., 73.5 not 0.735)
FMT_INT = '0'
FMT_RATIO = '0.00x'

GREEN_FONT = Font(size=10, color="006100")
RED_FONT = Font(size=10, color="9C0006")


# ── Public API ──

def export_to_excel(
    analysis: AnalysisResult,
    filepath,  # str (file path) or BytesIO buffer
    ingested: dict[str, NormalizedResult] | None = None,
    company_name: str = "",
):
    """
    Export analysis results to a formatted Excel workbook.

    filepath can be a string path or a BytesIO buffer.
    All math is pre-computed. This only renders numbers into cells.
    """
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Build tabs
    _write_summary(wb, analysis, ingested, company_name)

    if analysis.ebitda_bridges:
        _write_ebitda_bridges(wb, analysis.ebitda_bridges)

    if analysis.variance:
        _write_variance(wb, analysis.variance)

    if analysis.margins:
        _write_margins(wb, analysis.margins)

    if analysis.working_capital:
        _write_working_capital(wb, analysis.working_capital)

    if analysis.fcf:
        _write_fcf(wb, analysis.fcf)

    if analysis.ltm:
        _write_ltm(wb, analysis.ltm)

    if analysis.revenue_analytics:
        _write_revenue_analytics(wb, analysis.revenue_analytics)

    if analysis.trends:
        _write_trends(wb, analysis.trends)

    if ingested:
        _write_raw_data(wb, ingested)

    wb.save(filepath)


# ── Formatting helpers ──

def _header_row(ws, row: int, values: list[str], col_start: int = 1):
    """Write a blue header row."""
    for i, val in enumerate(values):
        cell = ws.cell(row=row, column=col_start + i, value=val)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _section_row(ws, row: int, label: str, num_cols: int = 1):
    """Write a light blue section header spanning columns."""
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = SECTION_FILL
        cell.border = THIN_BORDER
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL


def _data_cell(ws, row: int, col: int, value, fmt: str | None = None, bold: bool = False, color_sign: bool = False):
    """Write a data cell with optional formatting."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = BOLD_FONT if bold else DATA_FONT
    cell.border = THIN_BORDER

    if fmt:
        cell.number_format = fmt

    if color_sign and value is not None and isinstance(value, (int, float)):
        if value > 0:
            cell.font = GREEN_FONT if not bold else Font(bold=True, size=10, color="006100")
        elif value < 0:
            cell.font = RED_FONT if not bold else Font(bold=True, size=10, color="9C0006")

    return cell


def _label_cell(ws, row: int, col: int, value: str, bold: bool = True):
    """Write a label cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = LABEL_FONT if bold else DATA_FONT
    cell.border = THIN_BORDER
    return cell


def _set_col_widths(ws, widths: dict[str, int]):
    """Set column widths. Keys are column letters."""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _val(v, default=""):
    """Return value or default if None."""
    return v if v is not None else default


# ── Tab writers ──

def _write_summary(wb, analysis, ingested, company_name):
    ws = wb.create_sheet("Summary")
    _set_col_widths(ws, {"A": 25, "B": 30, "C": 20})

    r = 1
    _label_cell(ws, r, 1, "PE Value Creation Analysis", bold=True)
    ws.cell(row=r, column=1).font = Font(bold=True, size=14, color="1F4E79")
    r += 1
    _label_cell(ws, r, 1, "Company:")
    _data_cell(ws, r, 2, company_name or "N/A")
    r += 1
    _label_cell(ws, r, 1, "Analysis Date:")
    _data_cell(ws, r, 2, date.today().isoformat())
    r += 1
    _label_cell(ws, r, 1, "Modules Run:")
    _data_cell(ws, r, 2, ", ".join(analysis.modules_run))
    r += 2

    # Data sources
    if ingested:
        _section_row(ws, r, "Data Sources", 3)
        r += 1
        _header_row(ws, r, ["Document Type", "Rows", "Quality Score"])
        r += 1
        for doc_type, result in ingested.items():
            _label_cell(ws, r, 1, result.doc_type_name, bold=False)
            _data_cell(ws, r, 2, len(result.df))
            _data_cell(ws, r, 3, result.quality_score, fmt=FMT_PCT_RAW)
            r += 1
        r += 1

    # Key metrics snapshot
    if analysis.margins and analysis.margins.periods:
        m = analysis.margins.periods[-1]
        _section_row(ws, r, f"Key Metrics — {m.period}", 3)
        r += 1
        metrics = [
            ("Gross Margin", m.gross_margin_pct, "%"),
            ("EBITDA Margin", m.ebitda_margin_pct, "%"),
            ("Revenue Growth MoM", m.revenue_growth_mom, "%"),
            ("Revenue Growth YoY", m.revenue_growth_yoy, "%"),
            ("EBITDA Growth MoM", m.ebitda_growth_mom, "%"),
            ("OpEx % of Revenue", m.opex_pct_revenue, "%"),
        ]
        for label, val, unit in metrics:
            _label_cell(ws, r, 1, label, bold=False)
            if val is not None:
                _data_cell(ws, r, 2, f"{val:.1f}{unit}")
            else:
                _data_cell(ws, r, 2, "N/A")
            r += 1

    if analysis.warnings:
        r += 1
        _section_row(ws, r, "Warnings", 3)
        r += 1
        for w in analysis.warnings:
            _data_cell(ws, r, 1, w)
            r += 1

    ws.freeze_panes = "A2"


def _write_bridge(ws, bridge: EBITDABridge, start_row: int) -> int:
    """Write one EBITDA bridge. Returns the next available row."""
    r = start_row
    _section_row(ws, r, f"{bridge.label}  ({bridge.base_period} → {bridge.current_period})", 3)
    r += 1
    _header_row(ws, r, ["Component", "Amount ($)", ""])
    r += 1

    _label_cell(ws, r, 1, f"Starting EBITDA ({bridge.base_period})")
    _data_cell(ws, r, 2, bridge.base_ebitda, fmt=FMT_CURRENCY, bold=True)
    r += 1

    for comp in bridge.components:
        _label_cell(ws, r, 1, f"  {comp.name}", bold=False)
        _data_cell(ws, r, 2, comp.value, fmt=FMT_CURRENCY_SIGN, color_sign=True)
        r += 1

    _label_cell(ws, r, 1, f"Ending EBITDA ({bridge.current_period})")
    _data_cell(ws, r, 2, bridge.current_ebitda, fmt=FMT_CURRENCY, bold=True)
    r += 1

    _label_cell(ws, r, 1, "Total Change")
    _data_cell(ws, r, 2, bridge.total_change, fmt=FMT_CURRENCY_SIGN, bold=True, color_sign=True)
    r += 1

    _label_cell(ws, r, 1, "Verified", bold=False)
    _data_cell(ws, r, 2, "Yes" if bridge.is_verified else f"No (delta: {bridge.verification_delta:.2f})")
    r += 1

    return r + 1  # blank row after


def _write_ebitda_bridges(wb, bridges):
    ws = wb.create_sheet("EBITDA Bridges")
    _set_col_widths(ws, {"A": 30, "B": 18, "C": 12})

    r = 1
    if bridges.mom:
        r = _write_bridge(ws, bridges.mom, r)
    if bridges.vs_budget:
        r = _write_bridge(ws, bridges.vs_budget, r)
    if bridges.vs_prior_year:
        r = _write_bridge(ws, bridges.vs_prior_year, r)

    ws.freeze_panes = "A2"


def _write_variance(wb, variance):
    ws = wb.create_sheet("Variance Analysis")

    # Use latest period
    latest = variance.periods[-1]
    _set_col_widths(ws, {"A": 22, "B": 16, "C": 16, "D": 16, "E": 12, "F": 12})

    r = 1
    _section_row(ws, r, f"Variance Analysis — {latest.period}", 6)
    r += 1

    # vs Prior Month
    if latest.vs_prior_month:
        _header_row(ws, r, ["Line Item", "Actual ($)", "Prior Month ($)", "$ Change", "% Change", "Fav/Unfav"])
        r += 1
        for v in latest.vs_prior_month:
            _label_cell(ws, r, 1, v.line_item.replace("_", " ").title(), bold=False)
            _data_cell(ws, r, 2, v.actual, fmt=FMT_CURRENCY)
            _data_cell(ws, r, 3, v.comparator, fmt=FMT_CURRENCY)
            _data_cell(ws, r, 4, v.dollar_change, fmt=FMT_CURRENCY_SIGN, color_sign=True)
            _data_cell(ws, r, 5, f"{v.pct_change:+.1f}%" if v.pct_change is not None else "N/A")
            fav_text = "Favorable" if v.favorable == Favorability.FAVORABLE else ("Unfavorable" if v.favorable == Favorability.UNFAVORABLE else "—")
            cell = _data_cell(ws, r, 6, fav_text)
            if v.favorable == Favorability.FAVORABLE:
                cell.font = GREEN_FONT
            elif v.favorable == Favorability.UNFAVORABLE:
                cell.font = RED_FONT
            r += 1
        r += 1

    # vs Budget
    if latest.vs_budget:
        _section_row(ws, r, "vs Budget", 6)
        r += 1
        _header_row(ws, r, ["Line Item", "Actual ($)", "Budget ($)", "$ Change", "% Change", "Fav/Unfav"])
        r += 1
        for v in latest.vs_budget:
            _label_cell(ws, r, 1, v.line_item.replace("_", " ").title(), bold=False)
            _data_cell(ws, r, 2, v.actual, fmt=FMT_CURRENCY)
            _data_cell(ws, r, 3, v.comparator, fmt=FMT_CURRENCY)
            _data_cell(ws, r, 4, v.dollar_change, fmt=FMT_CURRENCY_SIGN, color_sign=True)
            _data_cell(ws, r, 5, f"{v.pct_change:+.1f}%" if v.pct_change is not None else "N/A")
            fav_text = "Favorable" if v.favorable == Favorability.FAVORABLE else ("Unfavorable" if v.favorable == Favorability.UNFAVORABLE else "—")
            cell = _data_cell(ws, r, 6, fav_text)
            if v.favorable == Favorability.FAVORABLE:
                cell.font = GREEN_FONT
            elif v.favorable == Favorability.UNFAVORABLE:
                cell.font = RED_FONT
            r += 1

    ws.freeze_panes = "A3"


def _write_margins(wb, margins):
    ws = wb.create_sheet("Margins & Growth")

    periods = margins.periods
    num_periods = len(periods)

    _set_col_widths(ws, {"A": 22})
    for i in range(num_periods):
        ws.column_dimensions[get_column_letter(i + 2)].width = 14

    r = 1
    # Header: Metric | Period1 | Period2 | ...
    header = ["Metric"] + [p.period for p in periods]
    _header_row(ws, r, header)
    r += 1

    # Rows — each metric
    metric_rows = [
        ("Gross Margin %", "gross_margin_pct", FMT_PCT_RAW),
        ("EBITDA Margin %", "ebitda_margin_pct", FMT_PCT_RAW),
        ("S&M % Revenue", "sm_pct_revenue", FMT_PCT_RAW),
        ("R&D % Revenue", "rd_pct_revenue", FMT_PCT_RAW),
        ("G&A % Revenue", "ga_pct_revenue", FMT_PCT_RAW),
        ("OpEx % Revenue", "opex_pct_revenue", FMT_PCT_RAW),
        ("Rev Growth MoM %", "revenue_growth_mom", FMT_PCT_RAW),
        ("Rev Growth YoY %", "revenue_growth_yoy", FMT_PCT_RAW),
        ("EBITDA Growth MoM %", "ebitda_growth_mom", FMT_PCT_RAW),
        ("EBITDA Growth YoY %", "ebitda_growth_yoy", FMT_PCT_RAW),
    ]

    for label, attr, fmt in metric_rows:
        _label_cell(ws, r, 1, label)
        for i, p in enumerate(periods):
            val = getattr(p, attr, None)
            is_growth = "growth" in attr.lower()
            _data_cell(ws, r, i + 2, _val(val), fmt=fmt, color_sign=is_growth)
        r += 1

    ws.freeze_panes = "B2"


def _write_ltm(wb, ltm):
    ws = wb.create_sheet("LTM & Rule of 40")
    _set_col_widths(ws, {"A": 28, "B": 20})

    r = 1
    _section_row(ws, r, f"Last Twelve Months — as of {ltm.as_of_period}", 2)
    r += 1
    _header_row(ws, r, ["Metric", "Value"])
    r += 1

    rows = [
        ("LTM Revenue", ltm.ltm_revenue, FMT_CURRENCY),
        ("LTM COGS", ltm.ltm_cogs, FMT_CURRENCY),
        ("LTM Gross Profit", ltm.ltm_gross_profit, FMT_CURRENCY),
        ("LTM EBITDA", ltm.ltm_ebitda, FMT_CURRENCY),
        ("LTM Gross Margin %", ltm.ltm_gross_margin_pct, FMT_PCT_RAW),
        ("LTM EBITDA Margin %", ltm.ltm_ebitda_margin_pct, FMT_PCT_RAW),
        ("LTM Revenue Growth YoY %", ltm.ltm_revenue_growth_yoy, FMT_PCT_RAW),
        ("Months Included", ltm.months_included, FMT_INT),
    ]

    for label, val, fmt in rows:
        _label_cell(ws, r, 1, label)
        _data_cell(ws, r, 2, _val(val), fmt=fmt)
        r += 1

    r += 1
    _section_row(ws, r, "SaaS Efficiency", 2)
    r += 1
    _label_cell(ws, r, 1, "Rule of 40")
    if ltm.rule_of_40 is not None:
        _data_cell(ws, r, 2, ltm.rule_of_40, fmt=FMT_PCT_RAW,
                   bold=True, color_sign=True)
    else:
        _data_cell(ws, r, 2, "N/A (need 13+ months)")
    r += 1
    _label_cell(ws, r, 1, "  = Revenue Growth % + EBITDA Margin %", bold=False)
    if ltm.ltm_revenue_growth_yoy is not None and ltm.ltm_ebitda_margin_pct is not None:
        _data_cell(ws, r, 2, f"{ltm.ltm_revenue_growth_yoy:.1f}% + {ltm.ltm_ebitda_margin_pct:.1f}%")
    r += 1
    _label_cell(ws, r, 1, "  Target: > 40", bold=False)

    ws.freeze_panes = "A3"


def _write_working_capital(wb, wc):
    ws = wb.create_sheet("Working Capital")

    periods = wc.periods
    _set_col_widths(ws, {"A": 22})
    for i in range(len(periods)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 16

    r = 1
    header = ["Metric"] + [p.period for p in periods]
    _header_row(ws, r, header)
    r += 1

    # DSO, DPO, DIO, CCC
    for label, attr, fmt in [
        ("DSO (days)", "dso", FMT_INT),
        ("DPO (days)", "dpo", FMT_INT),
        ("DIO (days)", "dio", FMT_INT),
        ("Cash Conversion Cycle", "ccc", FMT_INT),
        ("WC Change ($)", "wc_change", FMT_CURRENCY_SIGN),
        ("DSO Cash Impact ($)", "dso_cash_impact", FMT_CURRENCY_SIGN),
    ]:
        _label_cell(ws, r, 1, label)
        for i, p in enumerate(periods):
            val = getattr(p, attr, None)
            color = attr in ("wc_change", "dso_cash_impact")
            _data_cell(ws, r, i + 2, _val(val), fmt=fmt, color_sign=color)
        r += 1

    # AR Aging section
    has_aging = any(p.ar_aging for p in periods)
    if has_aging:
        r += 1
        _section_row(ws, r, "AR Aging Distribution (%)", len(periods) + 1)
        r += 1
        for label, attr in [
            ("Current (0-30)", "current_pct"),
            ("31-60 Days", "pct_31_60"),
            ("61-90 Days", "pct_61_90"),
            ("91-120 Days", "pct_91_120"),
            ("120+ Days", "over_120_pct"),
        ]:
            _label_cell(ws, r, 1, label, bold=False)
            for i, p in enumerate(periods):
                val = getattr(p.ar_aging, attr, None) if p.ar_aging else None
                _data_cell(ws, r, i + 2, _val(val), fmt=FMT_PCT_RAW)
            r += 1

    ws.freeze_panes = "B2"


def _write_fcf(wb, fcf):
    ws = wb.create_sheet("FCF & Leverage")

    periods = fcf.periods
    _set_col_widths(ws, {"A": 24})
    for i in range(len(periods)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 16

    r = 1
    header = ["Metric"] + [p.period for p in periods]
    _header_row(ws, r, header)
    r += 1

    for label, attr, fmt in [
        ("Free Cash Flow ($)", "free_cash_flow", FMT_CURRENCY),
        ("Cash Conversion Ratio", "cash_conversion_ratio", "0.00"),
        ("Net Debt ($)", "net_debt", FMT_CURRENCY_SIGN),
        ("LTM EBITDA ($)", "ltm_ebitda", FMT_CURRENCY),
        ("Net Debt / LTM EBITDA", "net_debt_to_ltm_ebitda", "0.0x"),
    ]:
        _label_cell(ws, r, 1, label)
        for i, p in enumerate(periods):
            val = getattr(p, attr, None)
            _data_cell(ws, r, i + 2, _val(val), fmt=fmt)
        r += 1

    ws.freeze_panes = "B2"


def _write_revenue_analytics(wb, ra):
    ws = wb.create_sheet("Revenue Analytics")
    _set_col_widths(ws, {"A": 22, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16})

    r = 1

    # Concentration
    if ra.concentration:
        _section_row(ws, r, "Revenue Concentration", 6)
        r += 1

        conc_periods = [c.period for c in ra.concentration]
        header = ["Metric"] + conc_periods
        _header_row(ws, r, header)
        r += 1

        for label, attr, fmt in [
            ("Top 1 %", "top1_pct", FMT_PCT_RAW),
            ("Top 5 %", "top5_pct", FMT_PCT_RAW),
            ("Top 10 %", "top10_pct", FMT_PCT_RAW),
            ("HHI Index", "herfindahl", "0.0000"),
            ("Count", "count", FMT_INT),
        ]:
            _label_cell(ws, r, 1, label)
            for i, c in enumerate(ra.concentration):
                val = getattr(c, attr, None)
                _data_cell(ws, r, i + 2, _val(val), fmt=fmt)
            r += 1
        r += 1

    # Price/Volume/Mix
    if ra.price_volume:
        _section_row(ws, r, "Price / Volume / Mix Decomposition", 6)
        r += 1

        pv_periods = [pv.period for pv in ra.price_volume]
        header = ["Component"] + pv_periods
        _header_row(ws, r, header)
        r += 1

        for label, attr in [
            ("Price Effect ($)", "price_effect"),
            ("Volume Effect ($)", "volume_effect"),
            ("Mix Effect ($)", "mix_effect"),
            ("Total Change ($)", "total_change"),
        ]:
            _label_cell(ws, r, 1, label)
            for i, pv in enumerate(ra.price_volume):
                val = getattr(pv, attr)
                _data_cell(ws, r, i + 2, val, fmt=FMT_CURRENCY_SIGN, color_sign=True)
            r += 1

        _label_cell(ws, r, 1, "Verified")
        for i, pv in enumerate(ra.price_volume):
            _data_cell(ws, r, i + 2, "Yes" if pv.is_verified else "No")
        r += 1
        r += 1

    # KPI Trends
    if ra.kpi_trends:
        _section_row(ws, r, "KPI Trends", 6)
        r += 1

        # Get all unique periods across all KPIs
        all_periods = set()
        for series in ra.kpi_trends.values():
            for period, _ in series:
                all_periods.add(period)
        sorted_periods = sorted(all_periods)

        header = ["KPI"] + sorted_periods
        _header_row(ws, r, header)
        r += 1

        KPI_FORMATS = {
            "cac": FMT_CURRENCY, "ltv": FMT_CURRENCY,
            "monthly_churn_rate": FMT_PCT_RAW, "net_revenue_retention": FMT_PCT_RAW,
            "gross_revenue_retention": FMT_PCT_RAW, "capacity_utilization": FMT_PCT_RAW,
            "total_headcount": FMT_INT, "nps_score": FMT_INT,
            "ltv_cac_ratio": "0.0",
        }

        for metric_name, series in ra.kpi_trends.items():
            display_name = metric_name.replace("_", " ").title()
            _label_cell(ws, r, 1, display_name, bold=False)
            period_vals = {p: v for p, v in series}
            fmt = KPI_FORMATS.get(metric_name, "0.0")
            for i, period in enumerate(sorted_periods):
                val = period_vals.get(period)
                _data_cell(ws, r, i + 2, _val(val), fmt=fmt)
            r += 1

    ws.freeze_panes = "B2"


def _write_trends(wb, trends):
    ws = wb.create_sheet("Trend Flags")
    _set_col_widths(ws, {"A": 12, "B": 25, "C": 22, "D": 14, "E": 12, "F": 60})

    r = 1
    _header_row(ws, r, ["Severity", "Metric", "Flag Type", "Value", "Period", "Detail"])
    r += 1

    severity_order = {Severity.CRITICAL: 0, Severity.WARNING: 1, Severity.INFO: 2}
    sorted_flags = sorted(trends.flags, key=lambda f: severity_order.get(f.severity, 99))

    for flag in sorted_flags:
        sev_text = flag.severity.value.upper()
        cell = _label_cell(ws, r, 1, sev_text)
        if flag.severity == Severity.CRITICAL:
            cell.font = Font(bold=True, size=10, color="9C0006")
        elif flag.severity == Severity.WARNING:
            cell.font = Font(bold=True, size=10, color="9C6500")
        else:
            cell.font = Font(bold=True, size=10, color="1F4E79")

        _data_cell(ws, r, 2, flag.metric.replace("_", " ").title())
        _data_cell(ws, r, 3, flag.flag_type.value.replace("_", " ").title())
        _data_cell(ws, r, 4, flag.current_value, fmt="0.0")
        _data_cell(ws, r, 5, flag.period)
        _data_cell(ws, r, 6, flag.detail)
        r += 1

    if not trends.flags:
        _data_cell(ws, r, 1, "No trend flags detected — all metrics within normal range.")

    ws.freeze_panes = "A2"


def _write_raw_data(wb, ingested):
    ws = wb.create_sheet("Raw Data")
    _set_col_widths(ws, {"A": 20})

    r = 1
    for doc_type, result in ingested.items():
        _section_row(ws, r, f"{result.doc_type_name} — {len(result.df)} rows", 10)
        r += 1

        df = result.df
        # Header
        for i, col in enumerate(df.columns):
            cell = ws.cell(row=r, column=i + 1, value=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(i + 1)].width = max(
                ws.column_dimensions[get_column_letter(i + 1)].width or 10, len(str(col)) + 4
            )
        r += 1

        # Data rows
        for _, row_data in df.iterrows():
            for i, col in enumerate(df.columns):
                val = row_data[col]
                if pd.isna(val):
                    _data_cell(ws, r, i + 1, "")
                elif isinstance(val, float) and abs(val) >= 100:
                    _data_cell(ws, r, i + 1, val, fmt=FMT_CURRENCY)
                else:
                    _data_cell(ws, r, i + 1, val)
            r += 1

        r += 2  # spacing between tables

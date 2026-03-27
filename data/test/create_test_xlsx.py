"""One-time script to generate the messy test Excel files."""
import openpyxl
from openpyxl.utils import get_column_letter


def create_manufacturing():
    """Multi-sheet Excel. P&L has COGS split into materials/labor/overhead."""
    wb = openpyxl.Workbook()

    # Dashboard sheet (decoy)
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash["A1"] = "Atlas Manufacturing — Q4 Dashboard"
    ws_dash["A3"] = "See P&L Detail for financial data"

    # Balance Sheet (decoy)
    ws_bs = wb.create_sheet("Balance Sheet")
    ws_bs["A1"] = "Balance Sheet — FY2025"

    # P&L Detail (the real one)
    ws = wb.create_sheet("P&L Detail")
    ws["A1"] = "Atlas Manufacturing"
    ws["A2"] = "Monthly Profit & Loss Statement"
    ws["A3"] = ""  # blank row

    headers = [
        "Period", "Net Sales", "Raw Materials", "Direct Labor",
        "Mfg Overhead", "Total COGS", "Gross Margin",
        "Selling Expenses", "R&D Costs", "Admin Expenses",
        "Total Operating Exp", "Operating Income"
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)

    data = [
        ("Apr-25", 5580000, 1674000, 1116000, 558000, 3348000, 2232000, 836000, 502000, 446000, 1784000, 448000),
        ("May-25", 5800000, 1740000, 1160000, 580000, 3480000, 2320000, 870000, 522000, 464000, 1856000, 464000),
        ("Jun-25", 6030000, 1809000, 1206000, 603000, 3618000, 2412000, 904000, 543000, 482000, 1929000, 483000),
        ("", None, None, None, None, None, None, None, None, None, None, None),  # Q2 separator
        ("Jul-25", 6270000, 1881000, 1254000, 627000, 3762000, 2508000, 940000, 564000, 502000, 2006000, 502000),
        ("Aug-25", 6520000, 1956000, 1304000, 652000, 3912000, 2608000, 978000, 587000, 522000, 2087000, 521000),
        ("Sep-25", 6780000, 2034000, 1356000, 678000, 4068000, 2712000, 1017000, 610000, 542000, 2169000, 543000),
        ("", None, None, None, None, None, None, None, None, None, None, None),  # Q3 separator
        ("Oct-25", 7050000, 2115000, 1410000, 705000, 4230000, 2820000, 1058000, 635000, 564000, 2257000, 563000),
        ("Nov-25", 7330000, 2199000, 1466000, 733000, 4398000, 2932000, 1100000, 660000, 586000, 2346000, 586000),
        ("Dec-25", 6900000, 2070000, 1380000, 690000, 4140000, 2760000, 1035000, 621000, 552000, 2208000, 552000),
        ("", None, None, None, None, None, None, None, None, None, None, None),  # Q4 separator
        ("Jan-26", 7200000, 2160000, 1440000, 720000, 4320000, 2880000, 1080000, 648000, 576000, 2304000, 576000),
        ("Feb-26", 7490000, 2247000, 1498000, 749000, 4494000, 2996000, 1124000, 674000, 599000, 2397000, 599000),
        ("Mar-26", 7800000, 2340000, 1560000, 780000, 4680000, 3120000, 1170000, 702000, 624000, 2496000, 624000),
    ]

    for row_idx, row_data in enumerate(data, 5):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Notes sheet (decoy)
    ws_notes = wb.create_sheet("Notes")
    ws_notes["A1"] = "Accounting notes and assumptions"

    wb.save("data/test/manufacturing_pl.xlsx")
    print("Created manufacturing_pl.xlsx")


def create_services():
    """Services company: combined SG&A, split revenue lines, budget columns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly P&L"

    headers = [
        "Month Ending", "Consulting Revenue", "Training Revenue",
        "Total Revenue", "Cost of Service Delivery", "Gross Profit",
        "SG&A", "Technology & Development", "EBITDA",
        "Budget Revenue", "Budget EBITDA"
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)

    data = [
        ("August 2025",   1680000, 420000, 2100000, 630000, 1470000, 735000, 294000, 441000, 2200000, 460000),
        ("September 2025", 1760000, 440000, 2200000, 660000, 1540000, 770000, 308000, 462000, 2250000, 470000),
        ("October 2025",  1840000, 460000, 2300000, 690000, 1610000, 805000, 322000, 483000, 2350000, 490000),
        ("November 2025", 1920000, 480000, 2400000, 720000, 1680000, 840000, 336000, 504000, 2400000, 510000),
        ("December 2025", 2000000, 500000, 2500000, 750000, 1750000, 875000, 350000, 525000, 2500000, 535000),
        ("January 2026",  2080000, 520000, 2600000, 780000, 1820000, 910000, 364000, 546000, 2650000, 560000),
        ("February 2026", 2160000, 540000, 2700000, 810000, 1890000, 945000, 378000, 567000, 2800000, 590000),
        ("March 2026",    2240000, 560000, 2800000, 840000, 1960000, 980000, 392000, 588000, 2900000, 610000),
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    wb.save("data/test/services_company.xlsx")
    print("Created services_company.xlsx")


def create_messy():
    """Worst case: headers on row 5, $000s, blanks, notes, abbreviations."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financials"

    ws["A1"] = "CONFIDENTIAL - Internal Use Only"
    # Row 2 blank
    ws["A3"] = "FY2025 Monthly P&L ($000s)"
    # Row 4 blank

    # Headers on row 5 with blank column B
    headers = ["Mo.", "", "Rev", "COS", "GP", "S&M", "Eng", "G&A", "Ttl OpEx", "EBITDA", "Bgt Rev", "Bgt EBITDA"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=5, column=i, value=h)

    # Data in $000s (so 1800 = $1,800,000)
    data = [
        ("Mar-25", None, 1800, 468, 1332, 630, 324, 216, 1170, 162, None, None),
        ("Apr-25", None, 1944, 505, 1439, 680, 350, 233, 1264, 175, None, None),
        ("May-25", None, 2099, 546, 1553, 735, 378, 252, 1364, 189, None, None),
        ("Jun-25", None, 2267, 589, 1678, 793, 408, 272, 1474, 204, None, None),
        ("Jul-25", None, 2448, 636, 1812, 857, 441, 294, 1591, 220, None, None),
        ("Aug-25", None, 2644, 687, 1957, 925, 476, 317, 1719, 238, None, None),
        ("Sep-25", None, 2749, 715, 2034, 962, 495, 330, 1787, 247, None, None),
        ("Oct-25", None, 2859, 743, 2116, 1001, 515, 343, 1858, 257, None, None),
        ("Nov-25", None, 2973, 773, 2200, 1041, 535, 357, 1932, 268, None, None),
        ("Dec-25", None, 3092, 804, 2288, 1082, 557, 371, 2009, 279, None, None),
        ("Jan-26", None, 3216, 835, 2381, 1126, 579, 385, 2090, 291, None, None),
        ("Feb-26", None, 3100, 837, 2263, 1116, 562, 378, 2056, 207, None, None),
        # blank row
        (None, None, None, None, None, None, None, None, None, None, None, None),
        # Current month with budget
        ("Mar-26", None, 3400, 901, 2499, 1190, 618, 402, 2210, 289, 3600, 340),
    ]

    for row_idx, row_data in enumerate(data, 6):
        for col_idx, val in enumerate(row_data, 1):
            if val is not None:
                ws.cell(row=row_idx, column=col_idx, value=val)

    # Full year total
    row_after = 6 + len(data)
    ws.cell(row=row_after, column=1, value="Full Year Total")

    # Notes
    ws.cell(row=row_after + 1, column=1, value="*Includes one-time restructuring charge in June")
    ws.cell(row=row_after + 2, column=1, value="*Engineering headcount increased from 12 to 18 in Q3")
    ws.cell(row=row_after + 3, column=1, value="*Revenue dip in Feb due to delayed enterprise deal")

    wb.save("data/test/messy_workbook.xlsx")
    print("Created messy_workbook.xlsx")


if __name__ == "__main__":
    create_manufacturing()
    create_services()
    create_messy()
    print("\nAll test files created.")
